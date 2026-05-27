"""Test local search names feature."""
import unittest
import json
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))


class TestLocalNamesPrompt(unittest.TestCase):

    def setUp(self):
        from location_analyzer import LocationAnalyzer
        self.analyzer = LocationAnalyzer.__new__(LocationAnalyzer)
        self.analyzer._log = lambda msg: None

    def test_prompt_requests_local_names_field(self):
        """Prompt must describe the local_names field."""
        locations = [
            {
                "name": "Hoboken",
                "type": "city",
                "admin_hierarchy": {
                    "parent_name": "Hudson County",
                    "level_4_name": "New Jersey",
                    "level_2_name": "United States",
                }
            }
        ]
        prompt = self.analyzer._prepare_batch_gpt_prompt(locations, 0)
        self.assertIn("local_names", prompt,
            "Prompt must ask for 'local_names' field")
        self.assertIn("Google Ads", prompt,
            "Prompt must mention Google Ads keywords context")

    def test_prompt_example_shows_local_names_array(self):
        """Example in prompt must show local_names as a JSON array."""
        locations = [{"name": "Hoboken", "type": "city", "admin_hierarchy": {}}]
        prompt = self.analyzer._prepare_batch_gpt_prompt(locations, 0)
        self.assertIn('"local_names"', prompt,
            "Example JSON in prompt must contain local_names key")


class TestLocalNamesParsing(unittest.TestCase):

    def setUp(self):
        from location_analyzer import LocationAnalyzer
        self.analyzer = LocationAnalyzer.__new__(LocationAnalyzer)
        self.analyzer._log = lambda msg: None
        self.analyzer.gpt_client = None
        self.analyzer.gpt_model = "gpt-4-turbo"
        self.analyzer.status_callback = lambda msg: None

    def test_gpt_parser_extracts_local_names(self):
        """_get_gpt_populations_batch must parse local_names from response."""
        from unittest.mock import MagicMock, patch

        mock_response = MagicMock()
        mock_response.choices = [MagicMock()]
        mock_response.choices[0].message.content = json.dumps([
            {
                "index": 0,
                "population": 52000,
                "confidence": "High",
                "local_names": ["Hoboken", "Hoboken NJ", "Mile Square City"]
            }
        ])

        locations = [{"name": "Hoboken", "type": "city", "admin_hierarchy": {}}]

        with patch.object(self.analyzer, 'gpt_client') as mock_client:
            mock_client.chat.completions.create.return_value = mock_response
            result = self.analyzer._get_gpt_populations_batch(locations, 0)

        self.assertIn(0, result)
        self.assertEqual(result[0]["population"], 52000)
        self.assertEqual(result[0]["local_names"], ["Hoboken", "Hoboken NJ", "Mile Square City"])

    def test_gpt_parser_defaults_local_names_to_official_name(self):
        """If local_names is missing from response, fall back to [official name]."""
        from unittest.mock import MagicMock, patch

        mock_response = MagicMock()
        mock_response.choices = [MagicMock()]
        mock_response.choices[0].message.content = json.dumps([
            {"index": 0, "population": 52000, "confidence": "High"}
        ])

        locations = [{"name": "Hoboken", "type": "city", "admin_hierarchy": {}}]

        with patch.object(self.analyzer, 'gpt_client') as mock_client:
            mock_client.chat.completions.create.return_value = mock_response
            result = self.analyzer._get_gpt_populations_batch(locations, 0)

        self.assertIn(0, result)
        self.assertEqual(result[0]["local_names"], ["Hoboken"],
            "Should fall back to official name when local_names missing")


class TestLocalNamesStorage(unittest.TestCase):

    def test_estimate_populations_initializes_local_names(self):
        """estimate_populations must initialize local_names on every location."""
        from location_analyzer import LocationAnalyzer
        analyzer = LocationAnalyzer.__new__(LocationAnalyzer)
        analyzer._log = lambda msg: None
        analyzer.use_gpt = False
        analyzer.use_openai = False
        analyzer.use_gemini_flag = False
        analyzer._ensure_ai_clients = lambda: None
        analyzer.chunk_size = 5

        locations = [
            {"name": "Hoboken", "type": "city", "admin_hierarchy": {}},
            {"name": "Weehawken", "type": "suburb", "admin_hierarchy": {}},
        ]
        result = analyzer.estimate_populations(locations)

        for loc in result:
            self.assertIn("local_names", loc,
                f"Location '{loc.get('name')}' must have 'local_names' after estimate_populations")
            self.assertIsInstance(loc["local_names"], list)

    def test_estimate_populations_stores_gpt_local_names(self):
        """estimate_populations must copy local_names from GPT batch result into each location."""
        from location_analyzer import LocationAnalyzer
        from unittest.mock import patch
        analyzer = LocationAnalyzer.__new__(LocationAnalyzer)
        analyzer._log = lambda msg: None
        analyzer.use_gpt = True
        analyzer.use_openai = True
        analyzer.use_gemini_flag = False
        analyzer._ensure_ai_clients = lambda: None
        analyzer.chunk_size = 5
        analyzer.status_callback = lambda msg: None
        analyzer.gpt_model = "gpt-4-turbo"

        locations = [{"name": "Hoboken", "type": "city", "admin_hierarchy": {}}]

        with patch.object(analyzer, '_get_gpt_populations_batch', return_value={
            0: {"population": 52000, "confidence": "High", "local_names": ["Hoboken", "Mile Square City"]}
        }):
            result = analyzer.estimate_populations(locations)

        self.assertEqual(result[0]["local_names"], ["Hoboken", "Mile Square City"])


class TestLocalNamesExcel(unittest.TestCase):

    def _make_test_locations(self):
        return [
            {
                "name": "Hoboken",
                "type": "city",
                "latitude": 40.744,
                "longitude": -74.032,
                "gpt_population": 52000,
                "gpt_confidence": "High",
                "gemini_population": 51000,
                "gemini_confidence": "High",
                "combined_population": 51500,
                "combined_confidence": "High",
                "local_names": ["Hoboken", "Mile Square City", "Hoboken NJ"],
                "admin_hierarchy": {},
            }
        ]

    def _get_sheet_headers(self, output_bytes, sheet_name):
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(output_bytes))
        ws = wb[sheet_name]
        return [cell.value for cell in ws[1]]

    def _get_cell_value(self, output_bytes, sheet_name, row, col_name):
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(output_bytes))
        ws = wb[sheet_name]
        headers = [cell.value for cell in ws[1]]
        col_idx = headers.index(col_name) + 1
        return ws.cell(row=row, column=col_idx).value

    def test_full_data_sheet_has_local_search_names_column(self):
        """Full Data sheet must include a 'Local Search Names' column."""
        from location_analyzer import LocationAnalyzer
        analyzer = LocationAnalyzer.__new__(LocationAnalyzer)
        analyzer._log = lambda msg: None
        analyzer.output_excel = "test_output.xlsx"

        output = analyzer.save_to_excel(self._make_test_locations())
        self.assertIsNotNone(output)
        output_bytes = output.read()

        headers = self._get_sheet_headers(output_bytes, "Full Data")
        self.assertIn("Local Search Names", headers,
            f"Full Data sheet must have 'Local Search Names' column, got: {headers}")

        cell_value = self._get_cell_value(output_bytes, "Full Data", 2, "Local Search Names")
        self.assertIn("Hoboken", cell_value,
            f"Local Search Names cell should contain 'Hoboken', got: {cell_value!r}")

    def test_clean_sheet_has_local_search_names_column(self):
        """Clean Data sheet must include a 'Local Search Names' column."""
        from location_analyzer import LocationAnalyzer
        analyzer = LocationAnalyzer.__new__(LocationAnalyzer)
        analyzer._log = lambda msg: None
        analyzer.output_excel = "test_output.xlsx"

        output = analyzer.save_to_excel(self._make_test_locations())
        self.assertIsNotNone(output)
        output_bytes = output.read()

        headers = self._get_sheet_headers(output_bytes, "Clean Data")
        self.assertIn("Local Search Names", headers,
            f"Clean Data sheet must have 'Local Search Names' column, got: {headers}")

    def test_local_names_are_comma_separated(self):
        """local_names list must be joined as comma-separated string in Excel."""
        from location_analyzer import LocationAnalyzer
        analyzer = LocationAnalyzer.__new__(LocationAnalyzer)
        analyzer._log = lambda msg: None
        analyzer.output_excel = "test_output.xlsx"

        output = analyzer.save_to_excel(self._make_test_locations())
        output_bytes = output.read()

        cell_value = self._get_cell_value(output_bytes, "Full Data", 2, "Local Search Names")
        self.assertIn(",", cell_value,
            f"Multiple local names should be comma-separated, got: {cell_value!r}")
        self.assertIn("Mile Square City", cell_value)
        self.assertIn("Hoboken NJ", cell_value)


if __name__ == "__main__":
    unittest.main()
